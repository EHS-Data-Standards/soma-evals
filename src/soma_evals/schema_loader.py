"""Load SOMA LinkML schema and Excel format as prompt context.

Provides two kinds of context for LLM prompts:
1. Schema context: LinkML class/slot definitions from YAML files
2. Format context: Excel sheet structure (sheet names + column headers)
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import yaml

_DEFAULT_SCHEMA_PATH = Path(__file__).parent.parent.parent.parent / "soma" / "src" / "soma" / "schema"
_DEFAULT_EXCEL_PATH = Path(__file__).parent.parent.parent.parent / "soma" / "project" / "excel" / "soma.xlsx"


def _resolve_schema_path() -> Path:
    """Resolve schema path from env var or default."""
    env = os.environ.get("SOMA_SCHEMA_PATH")
    return Path(env) if env else _DEFAULT_SCHEMA_PATH


def _resolve_excel_path() -> Path:
    """Resolve Excel path from env var or default."""
    env = os.environ.get("SOMA_EXCEL_PATH")
    return Path(env) if env else _DEFAULT_EXCEL_PATH


def load_schema_yaml(schema_path: Path | None = None) -> dict[str, Any]:
    """Load the main soma.yaml schema file and return parsed YAML."""
    path = schema_path or _resolve_schema_path()
    soma_yaml = path / "soma.yaml" if path.is_dir() else path
    if not soma_yaml.exists():
        msg = f"Schema file not found: {soma_yaml}"
        raise FileNotFoundError(msg)
    with open(soma_yaml) as f:
        data: dict[str, Any] = yaml.safe_load(f) or {}
    return data


def load_all_schema_yamls(schema_path: Path | None = None) -> dict[str, Any]:
    """Load all YAML files in the schema directory and merge classes/slots."""
    path = schema_path or _resolve_schema_path()
    if not path.is_dir():
        path = path.parent

    merged_classes: dict[str, Any] = {}
    merged_slots: dict[str, Any] = {}

    for yaml_file in sorted(path.glob("*.yaml")):
        if yaml_file.name == "README.md":
            continue
        with open(yaml_file) as f:
            data = yaml.safe_load(f) or {}
        if "classes" in data and isinstance(data["classes"], dict):
            merged_classes.update(data["classes"])
        if "slots" in data and isinstance(data["slots"], dict):
            merged_slots.update(data["slots"])

    return {"classes": merged_classes, "slots": merged_slots}


def schema_to_context_string(
    schema_path: Path | None = None,
    classes: list[str] | None = None,
) -> str:
    """Convert soma schema to a prompt-ready context string.

    Args:
        schema_path: path to schema directory (default: from env or ../soma)
        classes: if provided, only include these classes (and their slots).
                 If None, include all classes.

    Returns:
        A formatted string describing the schema classes and their slots.
    """
    data = load_all_schema_yamls(schema_path)
    all_classes = data.get("classes", {})
    all_slots = data.get("slots", {})

    if classes:
        filtered = {k: v for k, v in all_classes.items() if k in classes}
    else:
        filtered = all_classes

    lines: list[str] = ["# SOMA Schema (LinkML)", ""]

    for cls_name, cls_def in sorted(filtered.items()):
        if cls_def is None:
            continue
        desc = cls_def.get("description", "")
        is_a = cls_def.get("is_a", "")
        lines.append(f"## {cls_name}")
        if is_a:
            lines.append(f"  inherits: {is_a}")
        if desc:
            lines.append(f"  {desc}")
        lines.append("")

        # List slots for this class
        cls_slots = cls_def.get("slots", []) or []
        cls_attributes = cls_def.get("attributes", {}) or {}
        for slot_name in cls_slots:
            slot_def = all_slots.get(slot_name, {}) or {}
            slot_desc = slot_def.get("description", "")
            slot_range = slot_def.get("range", "string")
            multi = slot_def.get("multivalued", False)
            required = slot_def.get("required", False)
            tags: list[str] = []
            if required:
                tags.append("required")
            if multi:
                tags.append("multivalued")
            tag_str = f" [{', '.join(tags)}]" if tags else ""
            lines.append(f"  - **{slot_name}** ({slot_range}){tag_str}")
            if slot_desc:
                lines.append(f"    {slot_desc}")

        for attr_name, attr_def in cls_attributes.items():
            if attr_def is None:
                continue
            attr_desc = attr_def.get("description", "")
            attr_range = attr_def.get("range", "string")
            lines.append(f"  - **{attr_name}** ({attr_range})")
            if attr_desc:
                lines.append(f"    {attr_desc}")

        lines.append("")

    return "\n".join(lines)


def load_excel_headers(excel_path: Path | None = None) -> dict[str, list[str]]:
    """Load sheet names and column headers from the soma Excel serialization.

    Returns:
        dict mapping sheet name to list of column header strings.
    """
    import openpyxl

    path = excel_path or _resolve_excel_path()
    if not path.exists():
        msg = f"Excel file not found: {path}"
        raise FileNotFoundError(msg)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    headers: dict[str, list[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers[sheet_name] = [str(c) for c in row1 if c is not None]
    wb.close()
    return headers


def excel_to_format_string(
    excel_path: Path | None = None,
    sheets: list[str] | None = None,
) -> str:
    """Convert Excel structure to a prompt-ready format instruction string.

    Args:
        excel_path: path to soma.xlsx (default: from env or ../soma)
        sheets: if provided, only include these sheets. If None, include all.

    Returns:
        A formatted string describing the expected output format.
    """
    headers = load_excel_headers(excel_path)

    if sheets:
        headers = {k: v for k, v in headers.items() if k in sheets}

    lines: list[str] = [
        "# Expected Output Format",
        "",
        "Structure your output as data that could populate an Excel workbook",
        "with the following sheets and columns:",
        "",
    ]

    for sheet_name, cols in headers.items():
        lines.append(f"## Sheet: {sheet_name}")
        lines.append(f"  Columns: {', '.join(cols)}")
        lines.append("")

    lines.append("Output your extracted data as YAML, with top-level keys matching")
    lines.append("sheet names (in snake_case) and values as lists of records.")

    return "\n".join(lines)
